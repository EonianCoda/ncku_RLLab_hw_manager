import shutil
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QWidget
from PyQt5.QtCore import QThread, pyqtSignal
import PyQt5
# from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QFileDialog
from UI import Ui_Form
from hw_downloader import Hw_downloader
from collections import defaultdict
import os
import patoolib # for rar file
import pickle
import ftplib
from pathlib import Path
from collections import defaultdict
from functools import cmp_to_key
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
import re
from datetime import datetime, timedelta

FTP_HOWEWORK_PATH = "/Course/{}/Upload/Homework/"
# Course List
COURSE_FOLDERS = ['CvDl_2022_G', 'OpenCvDl_2022_Bs']
# CSV file contains the student's name and id
COURSE_STUDENT_LIST = ['./course_data/cvdl_students.csv', './course_data/opencvdl_students.csv']
TA_WHITE_LIST = ["./course_data/cvdl_TA_whitelist.csv", "./course_data/opencvdl_TA_whitelist.csv"]
# Log File
USER_LOG = "user.log"
LOGIN_KEY = "login.key"

DELAY_RATIO = [(timedelta(days=7), 0.5),(timedelta(days=365), 0)]

DOWNLOAD_FILE_ROOT = Path("./download")

##################### For Debug #########################
# if os.path.isdir(DOWNLOAD_FILE_ROOT):
#     shutil.rmtree(DOWNLOAD_FILE_ROOT)
##TODO cur_course == None修改
##TODO 新增下載log，避免多次重覆下載 
##TODO 新增開啟excel, 資料夾選項
##TODO List Directory timeout problem

##TODO GIF圖檔失真


login_success_styleSheet = """
QLabel
{	
	font: 75 20pt "微軟正黑體";
	font-weight:bold;
	color: green;
}
"""
def unzip(file_path : str, output_directory : str):
    """Unzip a zip file into directory
    Args:
        file_path: the path for zip file
        output_directory: the path for extract the zip file
    """
    os.makedirs(output_directory, exist_ok=True)
    output_directory = str(output_directory)
    file_path = str(file_path)
    patoolib.extract_archive(file_path, outdir=output_directory)

def clear_QTable(table, only_rows=False):
    cur_row = table.rowCount()
    for _ in range(cur_row):
        table.removeRow(0)

    # Clear col
    if only_rows == False:
        cur_col = table.columnCount()
        for _ in range(cur_col):
            table.removeColumn(0)

def compare_time(t1 : str, t2 :str):
    h1 = int(t1[0:2])
    h2 = int(t2[0:2])
    m1 = int(t1[3:5])
    m2 = int(t2[3:5])
    if h1 > h2 or (h1 == h2 and m1 > m2):
        return 1
    elif h1 < h2 or (h1 == h2 and m1 < m2):
        return -1
    else:
        return 0
def set_sheet_style(sheet, 
                    min_row : int,
                    min_col : int, 
                    max_row : int,
                    max_col : int,
                    border_style=None, 
                    fill_color=None):
    if border_style == None and fill_color == None:
        return

    
    for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            if border_style != None:
                cell.border = border_style
            if fill_color != None:
                cell.fill = fill_color
def set_table_item(table, text:str, row:int, col:int):
    item = QTableWidgetItem(str(text))
    item.setTextAlignment(PyQt5.QtCore.Qt.AlignCenter)
    table.setItem(row, col, item)
    table.resizeColumnsToContents()
class LoadingProgress(QtWidgets.QDialog):
    update_title_signal = pyqtSignal(str)
    update_content_signal = pyqtSignal(str)
    finished = pyqtSignal()
    def __init__(self, parent):
        super(LoadingProgress, self).__init__(parent)
        # self.setWindowFlags(self.windowFlags() ^ QtCore.Qt.WindowContextHelpButtonHint ^ QtCore.Qt.WindowCloseButtonHint)
        self.setWindowFlags(self.windowFlags() ^ QtCore.Qt.WindowContextHelpButtonHint)
        self.setWindowTitle("訊息")
        self.value = 0
        vbox = QtWidgets.QVBoxLayout(self)
        # Set gif movie
        self.movie_label = QtWidgets.QLabel()
        self.movie = QtGui.QMovie("./UI_imgs/loading.gif")
        self.movie_label.setMovie(self.movie)
        self.movie_label.setScaledContents(True)
        self.movie.start()
        self.progress_label = QtWidgets.QLabel()

        vbox.addWidget(self.movie_label)
        vbox.addWidget(self.progress_label)
        
        self.setLayout(vbox)
        self.setStyleSheet("background-color:white;")

        self.setup_signal()
    def setup_signal(self):
        self.finished.connect(self.close_progress)
        self.update_title_signal.connect(self.update_title)
        self.update_content_signal.connect(self.update_content)
    def close_progress(self):
        self.close()
    def update_title(self, title:str):
        self.setWindowTitle(title)
    def update_content(self, content:str):
        self.progress_label.setText(content)

class User_log(object):
    def __init__(self):
        self.delay_time_log = dict()
        self.download_log = dict()
    def add_delay_time_log(self, course : str, sub_hw : str, timestamp):
        key = "{}/{}".format(course, sub_hw)
        self.delay_time_log[key] = timestamp
    def get_delay_time_log(self, course : str, sub_hw : str):
        key = "{}/{}".format(course, sub_hw)
        return self.delay_time_log.get(key)

    def add_download_log(self, course:str, sub_hw:str, s_id:str, output_path:str):
        key = "{}/{}".format(course, sub_hw)
        files_dict = self.download_log.get(key)
        if files_dict == None:
            files_dict = dict()
            self.download_log[key] = files_dict

        files_dict[s_id] = output_path
    def check_download_exist(self, course: str, sub_hw: str, s_id:str, output_path: str):
        key = "{}/{}".format(course, sub_hw)
        files_dict = self.download_log.get(key)
        if files_dict == None:
            return False
        old_output_path = files_dict.get(s_id)
        if old_output_path == None or old_output_path != output_path:
            return False
        # The file exists
        return True

    def write_user_log(self):
        with open(USER_LOG, 'wb') as f:
            pickle.dump(self, f)
class Download_files_thread(QThread):
    def __init__(self, 
                loading_bar:LoadingProgress,
                downloader : Hw_downloader,
                download_s_ids : list,
                store_dirs : dict,
                course: str,
                hw : str,
                submitted_files : dict,
                user_log : User_log,
        ):
        """
        Attributes:
            loading_bar: a LoadingProgress object for signaling the current process
            downaloder: a Hw_downloader object for managing ftp
            store_dirs: a dict for pair{s_id: store_dir}, s_id mean student id and store_dir is a directory for storing file
            course: string for which course will be downloaded
            hw: string for which homework will be downloaded

            user_log: a User_log object for logging the file which be downloaded
        """
        super().__init__()
        self.loading_bar = loading_bar

        self.downloader = downloader
        self.download_s_ids = download_s_ids
        self.store_dirs = store_dirs
        self.course = course
        self.hw = hw
        self.submitted_files = submitted_files
        key = "{}/{}".format(self.course, self.hw)
        self.hw_folders = [k for k in self.submitted_files[key].keys()]
        if len(self.hw_folders) == 0:
            raise ValueError("The key {} does not exist in submitted_files".format(key))
        self.user_log = user_log
    def run(self):
        self.loading_bar.update_title_signal.emit("下載檔案中")
        
        download_files_path = []
        

        error_lines = ['錯誤檔案名稱, 錯誤類別, 錯誤訊息\n']
        reconnect_times = 0
        remove_list = []
        # Download Files
        
        key = "{}/{}".format(self.course, self.hw)
        for sub_hw in self.hw_folders:
            ftp_target_path = FTP_HOWEWORK_PATH.format(self.course) + sub_hw

            all_datas, _ = self.submitted_files[key][sub_hw]
            for idx, s_id in enumerate(self.download_s_ids):
                datas = all_datas.get(s_id)
                if datas == None:
                    error_lines.append("{},{},{}\n".format("{}/{}".format(sub_hw, s_id), "需手動處理", "學號{} 沒有繳交作業{}".format(s_id, sub_hw)))
                    continue
                file_name = datas['file_name']
                ftp_path = "{}/{}".format(ftp_target_path, file_name)
                output_path = os.path.join(self.store_dirs[s_id], s_id)
                os.makedirs(output_path, exist_ok=True)
                output_path = os.path.join(output_path, file_name)
                # Check if this file was downloaded before.
                file_exist =  self.user_log.check_download_exist(self.course, sub_hw, s_id, output_path)
                if file_exist:
                    continue

                file_size = datas.get('file_size')
                if file_size == None:
                    file_size = "??"
                self.loading_bar.update_content_signal.emit("正在下載{}:({}/{}) {}({:}MB)".format(sub_hw, 
                                                                                                    idx, 
                                                                                                    len(self.download_s_ids), 
                                                                                                    file_name, 
                                                                                                    file_size))
                try:
                    self.downloader.download_file(ftp_path, output_path)
                    reconnect_times = 0
                    download_files_path.append(output_path)
                    self.user_log.add_download_log(self.course, sub_hw, s_id, output_path)
                    self.user_log.write_user_log()
                # No such file
                except ftplib.error_perm as e:
                    error_lines.append("{},{},{}\n".format(file_name, "download", e.__str__().replace(",","，")))
                    remove_list.append(output_path)
                except UnicodeEncodeError as e:
                    error_lines.append("{},{},{}\n".format(file_name, "download", e.__str__().replace(",","，")))
                    remove_list.append(output_path)
                # Server or Out Pc error
                except ftplib.error_temp as e:
                    if reconnect_times == 1:
                        print(e)
                        raise SystemError("Some Fatal error")
                    reconnect_times += 1
                    self.downloader.reconnect()
                    error_lines.append("{},{},{}\n".format(file_name, "download", "Exception ftplib.error_temp"))
                except ConnectionAbortedError as e:
                    if reconnect_times == 1:
                        print(e)
                        raise SystemError("Some Fatal error")
                    reconnect_times += 1
                    self.downloader.reconnect()
                    error_lines.append("{},{},{}\n".format(file_name, "download", "ConnectionAbortedError"))

        file_path = DOWNLOAD_FILE_ROOT / self.course / "{}_group_download_error.csv".format(self.hw)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('\ufeff')
            f.writelines(error_lines)

        os.system("start EXCEL.EXE {}".format(file_path))
        self.loading_bar.finished.emit()

class Unzip_thread(QThread):
    def __init__(self, 
                loading_bar:LoadingProgress,
                root_dir: str,
        ):
        super().__init__()
        self.loading_bar = loading_bar
        self.root_dir = root_dir
    def run(self):
        self.loading_bar.update_title_signal.emit("解壓縮檔案中")
        error_lines = ['錯誤檔案名稱, 錯誤類別, 錯誤訊息\n']
        unzip_files = []
        for root, dirs, files in os.walk(self.root_dir):
            if len(files) != 0:
                for f in files:
                    if '.7z' in f[-3:].lower() or '.rar' in f[-4:].lower() or '.zip'[-4:].lower():
                        unzip_files.append(os.path.join(root, f))
        remove_list = []
        for idx, file in enumerate(unzip_files):
            root_dir = os.path.dirname(file)
            out_directory = os.path.basename(file)
            if ".7z" in out_directory:
                out_directory = out_directory[:-3]
            else:
                out_directory = out_directory[:-4]
            output_path = os.path.join(root_dir, out_directory)
            self.loading_bar.update_content_signal.emit("正在解壓縮({}/{}) {}".format(idx, len(unzip_files),os.path.basename(file)))
            try:
                unzip(file, output_path)
                os.remove(file)
            # unzip error
            except Exception as e:
                remove_list.append(output_path)
                error_lines.append("{},{},{}\n".format(os.path.basename(file), "unzip", e.__str__().replace(",","，")))
        # Remove directory which failed to unzip
        for f in remove_list:
            shutil.rmtree(f)

        if len(error_lines) != 1:
            file_path = os.path.join("./download_and_unzip_error.csv")
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\ufeff')
                f.writelines(error_lines)
            os.system("start EXCEL.EXE {}".format(file_path))
        self.loading_bar.finished.emit()


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        
        self.ui.course_selecter.addItems(COURSE_FOLDERS)
        self.ui.course_selecter.currentTextChanged.connect(self.change_hw_options)
        self.submitted_files = dict()
        self.error_files = None
        self.ftp_target_path = None
        self.download_thread = None
        self.unzip_thread = None
        self.group_datas = None
        self.group_timestamps = None
        # Set Dialogue box and loading progress bar
        self.dlg = QMessageBox(self)
        self.loading_bar = LoadingProgress(self)
        self.setup_control()
        self.ui.course_selecter.setEnabled(False)
        self.ui.hw_selecter.setEnabled(False)

        # Load TA white list
        self.TA_whitelist = defaultdict(list)
        for course, csv_file in zip(COURSE_FOLDERS, TA_WHITE_LIST):
            with open(csv_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            for line in lines:
                line = line.strip()
                if line != "":
                    self.TA_whitelist[course].append(line.upper())

        # Load Students csv file
        self.student_IDS = defaultdict(list)
        self.student_names = defaultdict(list)
        self.student_major = defaultdict(list)
        for course, csv_file in zip(COURSE_FOLDERS, COURSE_STUDENT_LIST):
            with open(csv_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            white_list = self.TA_whitelist[course]
            for line in lines[1:]:
                line = line.replace("\u3000",'').replace("\n",'')
                data = line.split(',')
                name = data[0]
                stdID = data[3].strip().upper()
                major = data[4]
                if len(stdID) != 9 or stdID in white_list:
                    continue
                self.student_names[course].append(name)
                self.student_IDS[course].append(stdID)
                self.student_major[course].append(major)

        # Set downloader
        self.downloader = Hw_downloader(student_IDS = self.student_IDS,
                                        student_names = self.student_names)

        # If account log exist, then read it and auto login
        if os.path.exists(LOGIN_KEY):
            with open(LOGIN_KEY, 'rb') as f:
                username, password, auto_login = pickle.load(f)
            self.ui.username.setText(username)
            self.ui.password.setText(password)
            self.ui.auto_login_checkBox.setChecked(auto_login)
            # Auto login
            if auto_login:
                self.login_ftp()
        self.user_log = User_log()
        if os.path.exists(USER_LOG):
            with open(USER_LOG, 'rb') as f:
                self.user_log = pickle.load(f)
        
        clear_QTable(self.ui.set_group_table)
        self.ui.hw_submitted_status_table.resizeColumnsToContents()

    def group_table_select_all(self, checked : bool, col : int, all_flag=False):
        # Select All
        if all_flag:
            max_row = self.ui.set_group_table.rowCount()
            max_col = self.ui.set_group_table.columnCount()
            for row in range(max_row):
                for col in range(max_col):
                    cell = self.ui.set_group_table.cellWidget(row, col)
                    if cell != None:
                        cell.setChecked(checked)
            return
        # Seletct one column
        max_row = self.ui.set_group_table.rowCount()
        for row in range(max_row):
            cell = self.ui.set_group_table.cellWidget(row, col)
            if cell != None:
                cell.setChecked(checked)
        
    def check_login(self) -> bool:
        """Check whether ftp object exists
        """
        if self.downloader.ftp == None:
            self.show_not_connect_message()
            return False
        return True
    
    def setup_control(self):
        self.ui.login_btn.clicked.connect(self.login_ftp)
        self.ui.search_files_btn.clicked.connect(self.search_files)
        self.ui.download_hw_btn.clicked.connect(self.download_files)
        self.ui.output_submitted_info_btn.clicked.connect(self.output_submitted_info)
        self.ui.set_delay_time.clicked.connect(self.set_delay_time)
        self.ui.login_and_store_btn.clicked.connect(self.login_and_store)

        self.ui.hw_selecter.currentTextChanged.connect(self. change_hw_options_for_hw_selecter)
        self.ui.course_selecter.currentTextChanged.connect(self.set_search_path)

        self.ui.read_demo_time_file_btn.clicked.connect(self.select_demo_file)

        self.ui.download_hw_from_selection_btn.clicked.connect(self.download_from_selection)
        self.ui.write_delay_to_setting_btn.clicked.connect(self.write_delay_to_setting)
        self.ui.select_and_unzip_files_btn.clicked.connect(self.select_and_unzip_files)

        self.ui.import_files_btn.clicked.connect(self.import_files)
    def change_hw_options_for_hw_selecter(self, value : str):
        self.ui.sub_hw_selecter_for_delay_selecter.clear()
        self.ui.sub_hw_selecter_for_delay_selecter.addItems(self.hw_folders[self.cur_course_name][self.cur_hw])
    def import_files(self):
        filename, filetype = QFileDialog.getOpenFileName(self,
                                                        caption="Open file",
                                                        directory="./",
                                                        filter="Csv Files (*.csv)")
        if filename == "":
            return
        with open(filename, "r", encoding='utf-8') as f:
            lines = f.readlines()

        course_name = self.cur_course_name
        hw = self.cur_hw
        key = "{}/{}".format(course_name, hw)
        self.submitted_files[key] = dict()
        header = lines[0]
        header = header.split(',')
        sub_hws = [header[i][:-2] for i in range(3, len(header), 3)]

        clear_QTable(self.ui.hw_submitted_status_table, only_rows=True)
        for sub_hw, col_idx in zip(sub_hws, range(3, len(header), 3)):
            hw_files = dict()
            delay_count = 0
            for line in lines[1:]:
                datas = line.split(',')
                s_id = datas[0]
                file_name = datas[col_idx]
                if file_name != "":
                    hw_files[s_id] = {'timestamp': datetime.strptime(datas[col_idx + 1], "%Y/%m/%d %H:%M"),
                                    'file_name': file_name,
                                    'chinese_name': datas[1]}
                delay = datas[col_idx + 2]
                if delay.strip() != "":
                    delay_count += 1
            cur_row = self.ui.hw_submitted_status_table.rowCount()
            self.ui.hw_submitted_status_table.insertRow(cur_row)

            set_table_item(self.ui.hw_submitted_status_table, len(hw_files), cur_row, 0)
            set_table_item(self.ui.hw_submitted_status_table, 0, cur_row, 1)
            set_table_item(self.ui.hw_submitted_status_table, len(hw_files), cur_row, 2)
            set_table_item(self.ui.hw_submitted_status_table, self.cur_course_num_students - len(hw_files), cur_row, 3)
            set_table_item(self.ui.hw_submitted_status_table, delay_count, cur_row, 4)

            self.submitted_files[key][sub_hw] = (hw_files, [])
 
        self.ui.hw_submitted_status_table.setVerticalHeaderLabels(sub_hws)
        self.ui.hw_submitted_status_table.resizeColumnsToContents()

        # Set UI
        self.ui.output_submitted_info_btn.setEnabled(True)
        self.ui.download_hw_btn.setEnabled(True)
        self.ui.set_delay_time.setEnabled(True)
        self.ui.write_delay_to_setting_btn.setEnabled(True)
        self.ui.read_demo_time_file_btn.setEnabled(True)
        # Set delay time from user log
        cur_hw = self.cur_hw
        hw_folders = self.hw_folders[self.cur_course_name][cur_hw]
        for i, sub_hw in enumerate(hw_folders):
            timestamp = self.user_log.get_delay_time_log(self.cur_course_name, sub_hw)
            if timestamp != None:
                set_table_item(self.ui.hw_submitted_status_table, timestamp, i, 6)

    def select_and_unzip_files(self):
        if self.download_thread != None and self.download_thread.isFinished() == False:
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("請不要按兩次解壓縮，解壓縮還未完成")
            self.dlg.exec()
            return
        

        dir_name = QFileDialog.getExistingDirectory(self,
                                                    "Select Directory",
                                                    "./")
        if dir_name == "":
            return
        
        self.loading_bar.show()
        self.unzip_thread = Unzip_thread(self.loading_bar, dir_name)
        self.unzip_thread.start()
    def write_delay_to_setting(self):
        num_row = self.ui.hw_submitted_status_table.rowCount()
        cur_course = self.cur_course_name
        cur_hw = self.cur_hw
        hw_folders = self.hw_folders[cur_course][cur_hw]
        for sub_hw_idx in range(num_row):
            cell = self.ui.hw_submitted_status_table.item(sub_hw_idx, 6)
            if cell != None:
                self.user_log.add_delay_time_log(cur_course, hw_folders[sub_hw_idx], cell.text())
                
        self.user_log.write_user_log()  
                # timestamp = datetime.strptime(cell.text(), "%Y/%m/%d %H:%M")    
    def download_from_selection(self):
        if not self.check_login() or self.submitted_files == None:
            return
        if self.download_thread != None and self.download_thread.isFinished() == False:
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("請不要按兩次下載，下載還未完成")
            self.dlg.exec()
            if self.loading_bar.isVisible() == False:
                self.loading_bar.show()
            return
        
        root_dir = DOWNLOAD_FILE_ROOT / self.cur_course_name / "group_download" / self.cur_hw
        os.makedirs(root_dir, exist_ok=True)
        used_s_ids = []
        store_dirs = dict()
        max_row = self.ui.set_group_table.rowCount()
        max_col = self.ui.set_group_table.columnCount()
        for row in range(1, max_row):
            for col in range(1, max_col):
                cell = self.ui.set_group_table.cellWidget(row, col)
                if cell != None and cell.isChecked():
                    for s_id in self.group_datas[row - 1][col - 1]:
                        used_s_ids.append(s_id)
                        store_dirs[s_id] = os.path.join(root_dir, self.group_timestamps[row - 1].replace(':',''), "{}組".format(chr(65 + col - 1)))

        self.loading_bar.show()
        self.download_thread = Download_files_thread(loading_bar=self.loading_bar, 
                                                    downloader=self.downloader,
                                                    download_s_ids = used_s_ids,
                                                    store_dirs = store_dirs,
                                                    course = self.cur_course_name,
                                                    hw = self.cur_hw,
                                                    submitted_files = self.submitted_files,
                                                    user_log = self.user_log)
                                    
        self.download_thread.start()                               
    def select_demo_file(self):
        clear_QTable(self.ui.set_group_table)
        filename, filetype = QFileDialog.getOpenFileName(self,
                                                        caption="Open file",
                                                        directory="./",
                                                        filter="Csv Files (*.csv)")
        
        if filename == "":
            return
        if self.ui.am_num_TA_input.text() == "" or self.ui.pm_num_TA_input.text() == "":
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("請先填入上/下午分組數")
            self.dlg.exec()
            return
        num_TA = [int(self.ui.am_num_TA_input.text()), int(self.ui.pm_num_TA_input.text())]
        self.process_group_data(filename, num_TA)
    
    def login_ftp(self):
        # username or password is empty
        if self.ui.username.text() == "" or self.ui.password.text() == "":
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("帳號或密碼不得為空")
            self.dlg.exec()
            return False
        success = self.downloader.connect(self.ui.username.text(), self.ui.password.text())
        
        # Fail login
        if not success:
            self.dlg.setWindowTitle("登入失敗")
            self.dlg.setText("請重新登入")
            self.dlg.exec()
            return False
        self.search_folders()
        # Show dialogue
        self.dlg.setWindowTitle("登入成功")
        self.dlg.setText("歡迎使用")
        self.dlg.exec()
        # Update UI
        self.ui.login_status_label.setStyleSheet(login_success_styleSheet) 
        self.ui.login_status_label.setText("登入狀態: 已登入")
        self.ui.course_selecter.setEnabled(True)
        self.ui.hw_selecter.setEnabled(True)
        return True
    def login_and_store(self):
        success = self.login_ftp()
        # Store password
        if success:
            username = self.ui.username.text()
            password = self.ui.password.text()
            auto_login = self.ui.auto_login_checkBox.isChecked()
            with open("login.key", 'wb') as f:
                pickle.dump((username, password, auto_login), f)


    def set_delay_time(self):
        delay_time = self.ui.delay_time_set.dateTime().toPyDateTime().strftime("%Y/%m/%d %H:%M")
        idx = self.ui.sub_hw_selecter_for_delay_selecter.currentIndex()
        set_table_item(self.ui.hw_submitted_status_table, delay_time, idx, 6)
    def output_submitted_info(self):
        if not self.check_login() or self.submitted_files == None:
            return
        
        cur_course = self.cur_course_name
        cur_hw = self.cur_hw
        key = "{}/{}".format(cur_course, cur_hw)
        student_ids = self.student_IDS[cur_course]
        student_names = self.student_names[cur_course]
        student_major = self.student_major[cur_course]
        sorted_idx = sorted(range(len(student_ids)), key=lambda k : student_ids[k])

        cur_hw_folders = self.hw_folders[cur_course][cur_hw]
        contents = []
        fail_contents = ["子作業,檔案名稱,錯誤訊息\n"]
        header = "學號,姓名,科系"


        for idx in sorted_idx:
            s_id = student_ids[idx]
            name = student_names[idx]
            major = student_major[idx]
            contents.append("{},{},{}".format(s_id.upper(), name, major))

        for sub_hw_idx, sub_hw in enumerate(cur_hw_folders):
            header += ",{0}名稱,{0}上傳時間,{0}遲交分數比例".format(sub_hw)
            success_files, fail_files = self.submitted_files[key][sub_hw]

            delay_count = 0
            # Get Delay time
            delay_time = self.ui.hw_submitted_status_table.item(sub_hw_idx, 6)
            if delay_time != None:
                delay_time = datetime.strptime(delay_time.text(), "%Y/%m/%d %H:%M")
            # Process Success Files
            for line_idx, idx in enumerate(sorted_idx):
                s_id = student_ids[idx]
                if success_files.get(s_id) == None:
                    contents[line_idx] += ",,,"
                    continue
                data = success_files[s_id]

                delay_ratio = ""

                if delay_time != None and data['timestamp'] > delay_time:
                    for delta, ratio in DELAY_RATIO:
                        if data['timestamp'] - delay_time <= delta:
                            delay_ratio = ratio
                            break
                    delay_count += 1

                # timestamp = 
                contents[line_idx] += ",{},{},{}".format(data['file_name'],
                                                            data['timestamp'].strftime("%Y/%m/%d %H:%M"),
                                                            delay_ratio)
            set_table_item(self.ui.hw_submitted_status_table, delay_count, sub_hw_idx, 4)

            # Process Fail Files
            for (file_name, error_msg) in fail_files:
                fail_contents.append("{},{},{}\n".format(sub_hw, file_name, error_msg))
        csv_file_name = "{}.csv".format(self.cur_hw)
        output_path = DOWNLOAD_FILE_ROOT / self.cur_course_name
        os.makedirs(output_path, exist_ok=True)
        # Write log
        header += '\n'
        contents = [line + '\n' for line in contents]
        with open(output_path / csv_file_name, 'w', encoding='utf-8') as f:
            f.write('\ufeff')
            f.write(header)
            f.writelines(contents)

        fail_files_log = "{}_submit_error.csv".format(self.cur_hw)
        with open(output_path / fail_files_log, 'w', encoding='utf-8') as f:
            f.write('\ufeff')
            f.writelines(fail_contents)
        os.system("start EXCEL.EXE {}".format(str(output_path / csv_file_name)))
        if len(fail_contents) != 1: 
            os.system("start EXCEL.EXE {}".format(str(output_path / fail_files_log)))
        
    def download_files(self):
        if not self.check_login() or self.submitted_files == None:
            return
        if self.download_thread != None and self.download_thread.isFinished() == False:
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("請不要按兩次下載，下載還未完成")
            self.dlg.exec()
            return

        self.loading_bar.show()

        root_dir = DOWNLOAD_FILE_ROOT / self.cur_course_name / "normal_download" / self.cur_hw
        os.makedirs(root_dir, exist_ok=True)

        student_ids = self.student_IDS[self.cur_course_name][:3]
        store_dirs = {s_id:root_dir for s_id in student_ids}
        
        self.download_thread = Download_files_thread(loading_bar=self.loading_bar, 
                                                    downloader=self.downloader,
                                                    download_s_ids = student_ids,
                                                    store_dirs = store_dirs,
                                                    course = self.cur_course_name,
                                                    hw = self.cur_hw,
                                                    submitted_files = self.submitted_files,
                                                    user_log = self.user_log)
                                    
        self.download_thread.start()                                   
 
    def show_not_connect_message(self):
        self.dlg.setWindowTitle("錯誤")
        self.dlg.setText("請先登入後再執行此功能")
        self.dlg.exec()
    def search_files(self):
        if not self.check_login():
            return
        
        cur_course = self.cur_course_name
        cur_hw = self.cur_hw
        clear_QTable(self.ui.hw_submitted_status_table, only_rows=True)
        hw_folders = self.hw_folders[cur_course][self.cur_hw]
        key = "{}/{}".format(cur_course, cur_hw)
        self.submitted_files[key] = dict()
        for sub_hw in hw_folders:
            path = FTP_HOWEWORK_PATH.format(cur_course) + sub_hw
            success_files, success_count, fail_files = self.downloader.list_hw_files(path, cur_course)
            error_count = len(fail_files)
            cur_row = self.ui.hw_submitted_status_table.rowCount()
            self.ui.hw_submitted_status_table.insertRow(cur_row)

            # Success count, Fail count, num_submitted_student, num_no-submitted_student, delay_count, total_file_size
            set_table_item(self.ui.hw_submitted_status_table, success_count, cur_row, 0)
            set_table_item(self.ui.hw_submitted_status_table, error_count, cur_row, 1)
            set_table_item(self.ui.hw_submitted_status_table, len(success_files), cur_row, 2)
            set_table_item(self.ui.hw_submitted_status_table, self.cur_course_num_students - len(success_files), cur_row, 3)
            
            self.submitted_files[key][sub_hw] = (success_files, fail_files)


        self.ui.hw_submitted_status_table.setVerticalHeaderLabels(hw_folders)
        self.ui.hw_submitted_status_table.resizeColumnsToContents()

        # Set UI
        self.ui.output_submitted_info_btn.setEnabled(True)
        self.ui.download_hw_btn.setEnabled(True)
        self.ui.set_delay_time.setEnabled(True)
        self.ui.write_delay_to_setting_btn.setEnabled(True)
        self.ui.read_demo_time_file_btn.setEnabled(True)
        # Set delay time from user log
        cur_hw = self.cur_hw
        hw_folders = self.hw_folders[cur_course][cur_hw]
        for i, sub_hw in enumerate(hw_folders):
            timestamp = self.user_log.get_delay_time_log(cur_course, sub_hw)
            if timestamp != None:
                set_table_item(self.ui.hw_submitted_status_table, timestamp, i, 6)

    def set_search_path(self):
        cur_course = self.cur_course_name
        self.ui.search_files_btn.setEnabled(True)
        self.ui.import_files_btn.setEnabled(True)
        self.ui.num_student_label.setText("修課人數:{}".format(len(self.student_names[cur_course])))
        
    def process_group_data(self, group_csv_file:list, num_TA:tuple):
        """Process group data
        Args:
            group_csv_file: the path of group csv file
            num_TA: a tuple contains two value, first value means the number of TA in the morning, Second value means the number of TA in the afternonn
        """
        def create_cb_item():
            cb = QtWidgets.QCheckBox(parent=self.ui.set_group_table)
            cb.setStyleSheet("margin-left:10%; margin-right:10%;")
            return cb
        with open(group_csv_file,'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Set Table Header
        max_TA_num = max(num_TA)
        mark = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        self.ui.set_group_table.insertColumn(0)
        header = ['時段']
        for m in mark[:max_TA_num]:
            self.ui.set_group_table.insertColumn(0)
            header.append('{}組'.format(m))
        self.ui.set_group_table.setHorizontalHeaderLabels(header)

        # Set select all
        self.ui.set_group_table.insertRow(0)
        cb = create_cb_item()
        cb.clicked.connect(lambda checked: self.group_table_select_all(checked, 0, all_flag=True))
        self.ui.set_group_table.setCellWidget(0, 0, cb)
        
        # Set Column
        for col in range(1, max_TA_num + 1):
            cb = create_cb_item()
            cb.clicked.connect(lambda checked, col=col: self.group_table_select_all(checked, col))
            self.ui.set_group_table.setCellWidget(0, col, cb)


        num_TA_per_time = []
        num_student = 0


        student_name = dict()
        datas = defaultdict(list)
        all_s_ids = []
        for line in lines[1:]:
            data = line.split(',')
            s_id = data[2]
            timestamp = data[4]
            # Invalid Student id
            if len(s_id) != 9:
                continue
            if s_id not in self.TA_whitelist[self.cur_course_name]:
                all_s_ids.append(s_id)
            # No Selection
            if timestamp == "":
                continue
            
            # Store student_name
            student_name[s_id] = data[1]


            num_student += 1
            datas[timestamp].append(s_id)

        if len(all_s_ids) != self.cur_course_num_students:
            self.dlg.setWindowTitle("警告")
            self.dlg.setText("課程學生名單過舊 或 時間選擇名單過舊, 課程學生名稱:{}人, 時間選擇名單:{}人".format(self.cur_course_num_students, len(all_s_ids)))
            self.dlg.exec()
            clear_QTable(self.ui.set_group_table)
            return
        sorted_times = sorted(datas.keys(), key=cmp_to_key(compare_time))

        for time_key in sorted_times:
            s_ids = datas[time_key]
            s_ids = sorted(s_ids)
            datas[time_key] = s_ids

            hour = int(time_key[:2])
            # Afternoon
            if hour >= 13:
                num_TA_per_time.append(num_TA[1])
            else:
                num_TA_per_time.append(num_TA[0])

            # Update table
            cur_row = self.ui.set_group_table.rowCount()
            self.ui.set_group_table.insertRow(cur_row)
            # Add timestamp
            item = QTableWidgetItem(time_key)
            item.setTextAlignment(PyQt5.QtCore.Qt.AlignCenter)
            self.ui.set_group_table.setItem(cur_row, 0, item)

            for col in range(1, num_TA_per_time[-1] + 1):
                cb = create_cb_item()
                self.ui.set_group_table.setCellWidget(cur_row, col, cb)
        vertical_header = ['全選'] + [''] * len(sorted_times)
        self.ui.set_group_table.setVerticalHeaderLabels(vertical_header)

        self.ui.set_group_table.resizeColumnsToContents()

        # Style
        middle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        workbook = openpyxl.Workbook()
        sheet = workbook.worksheets[0]
        # Add header
        sheet.append(['時間','分組','姓名','學號'])
        set_sheet_style(sheet,
                        min_row=1,
                        max_row=1,
                        min_col=1,
                        max_col=50,
                        border_style=Border(bottom=Side(style='medium')))

        ### Set Timestamp ###
        cur_row = 2
        for time_key in sorted_times:
            # Set Time stamp
            timestamp = "{}\n~\n{}".format(time_key[0:5],time_key[-5:])
            cell = sheet.cell(row=cur_row, 
                            column=1, 
                            value = timestamp)
            # Set Style
            cell.font = Font(size=12, bold=True)
            cell.alignment = middle_alignment
            cell.border = Border(right=Side(style='thin'))
            row = cur_row + len(datas[time_key]) - 1
            set_sheet_style(sheet,
                        min_row=row,
                        max_row=row,
                        min_col=1,
                        max_col=50,
                        border_style=Border(bottom=Side(style='medium')))
            # Merge Cell
            sheet.merge_cells(start_row=cur_row,
                                end_row=cur_row + len(datas[time_key]) - 1,
                                start_column=1,
                                end_column=1)
            cur_row += len(datas[time_key])
            
        ### Set group and student ###
        cur_row = 2
        self.group_datas = []
        self.group_timestamps = sorted_times.copy()
        for row_idx, (time_key, num_group) in enumerate(zip(sorted_times, num_TA_per_time)):
            time_group_data = datas[time_key]
            self.group_datas.append([])
            # Compute number for each group
            per_num = len(time_group_data) // num_group
            per_student_group = [per_num] * num_group
            for i in range(len(time_group_data) - per_num * num_group):
                per_student_group[i] += 1
            
            cur_color_idx = 0
            cur_student_idx = 0
            for mark_idx, num in enumerate(per_student_group):
                cell = self.ui.set_group_table.cellWidget(row_idx + 1, mark_idx + 1)
                if cell != None:
                    cell.setText("{}人".format(num))
                if num == 0:
                    continue
                group_name = "{}組".format(mark[mark_idx])
                
                cell = sheet.cell(row=cur_row, 
                                    column=2, 
                                    value = group_name)
                # Set Style
                cell.font = Font(bold=True)
                cell.alignment =middle_alignment
                color_index = (45 + cur_color_idx) % 64
                # Color index = 0 is black
                if color_index == 0:
                    color_index = 1
                    cur_color_idx += 1
                # cell.fill = PatternFill('solid', Color(index=color_index))
                set_sheet_style(sheet,
                                min_row=cur_row,
                                max_row=cur_row + num - 1,
                                min_col=2,
                                max_col=2,
                                fill_color=PatternFill('solid', Color(index=color_index)))
                # Merge Cell
                sheet.merge_cells(start_row=cur_row,
                                end_row=cur_row + num - 1,
                                start_column=2,
                                end_column=2)
                if mark_idx != len(per_student_group) - 1:
                    set_sheet_style(sheet,
                                    min_row=cur_row + num - 1,
                                    max_row=cur_row + num - 1,
                                    min_col=2,
                                    max_col=50,
                                    border_style=Border(bottom=Side(style='dashed')))
                self.group_datas[-1].append([])
                # Student ids
                for i, s_id in enumerate(time_group_data[cur_student_idx : cur_student_idx+num]):
                    name = student_name[s_id]
                    sheet.cell(row=cur_row + i, column=3, value = name)
                    sheet.cell(row=cur_row + i, column=4, value = s_id)
                    self.group_datas[-1][-1].append(s_id)

                cur_student_idx += num
                cur_row += num
                cur_color_idx += 1

        # Set UI
        self.ui.download_hw_from_selection_btn.setEnabled(True)
        self.ui.set_group_table.resizeColumnsToContents()
        # Save File
        workbook.save('test.xlsx')
    def change_hw_options(self, value : str):
        self.ui.hw_selecter.clear()
        hws = [key for key in self.hw_folders[self.cur_course_name].keys()]
        self.ui.hw_selecter.addItems(hws)
        self.ui.sub_hw_selecter_for_delay_selecter.clear()
        self.ui.sub_hw_selecter_for_delay_selecter.addItems(self.hw_folders[self.cur_course_name][self.cur_hw])
    def search_folders(self):
        """Search homework folders
        """
        def compare_hw_folder(t1 : str, t2 :str):
            t1 = t1.split("_")
            t2 = t2.split("_")
            if len(t1) > 2 or len(t2) > 2:
                raise ValueError("The name of hw folder should be 'hw{hw_number}_{sub_homework}', hw_number should be a integer and sub_homework should be a str contain number, for example: '01','1','2'.")
            if len(t1) == 1:
                t1 = 0
            else:
                t1 = int(t1[1])
            if len(t2) == 1:
                t2 = 0
            else:
                t2 = int(t2[1])

            if t1 > t2:
                return 1
            elif t1 < t2:
                return -1
            else:
                return 0
        if self.downloader.ftp == None:
            self.show_not_connect_message()
            return
        
        self.hw_folders = dict()
        matcher = re.compile(u'hw[\d]')

        for course in COURSE_FOLDERS:
            hw_folders = self.downloader.list_folders(FTP_HOWEWORK_PATH.format(course))
            for name in hw_folders:
                match = matcher.search(name.lower())
                if match == None: # No version
                    raise ValueError("Homework folder name in FTP is invaild.Please Check it!")
                hw = match.group()
                if self.hw_folders.get(course) == None:
                    self.hw_folders[course] = defaultdict(list)

                self.hw_folders[course][hw].append(name)
        
        for course in COURSE_FOLDERS:
            for hw, hw_folders in self.hw_folders[course].items():
                sorted_hw_folders = sorted(hw_folders, key=cmp_to_key(compare_hw_folder))
                self.hw_folders[course][hw] = sorted_hw_folders
        # Set current select
        self.change_hw_options(None)


    def closeEvent(self, event) -> None:
        # Clost ftp connectoin
        if self.downloader.ftp != None:
            self.downloader.ftp.quit()
            print("結束連線")

    @property
    def cur_course_name(self):
        return self.ui.course_selecter.currentText()
    @property
    def cur_course_num_students(self):
        return len(self.student_names[self.cur_course_name])
    @property
    def cur_hw(self):
        return self.ui.hw_selecter.currentText()